#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""質譜 × 食品分析（PCA / PLS-DA）配套互動題庫產生器。
輸出 Kahoot(.xlsx) + Wordwall(.csv) + Gimkit(.csv) 到 ./game_imports/"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "game_imports")
os.makedirs(OUT, exist_ok=True)

# ── Kahoot：兩個遊戲（免費方案 ≤10 題）。[題目,A,B,C,D,秒,正解(1-4)] ──
KAHOOT_PCA = [
    ["PCA 屬於哪一種機器學習？", "監督式（需要標籤）", "非監督式（不需要標籤）", "強化學習", "迴歸預測", 20, 2],
    ["GC-MS 質譜分析後，資料通常整理成什麼？", "樣本 × 代謝物的數值矩陣", "一張照片", "一段影片", "單一數值", 20, 1],
    ["GC-MS 峰高在分析前先取對數（log）的目的？", "讓小峰消失", "增加樣本數", "壓縮跨數量級的動態範圍", "沒有作用", 20, 3],
    ["autoscale（自體尺度化）的目的是？", "讓高含量代謝物主導", "讓每個代謝物等權重、公平貢獻", "把資料全部歸零", "增加維度", 20, 2],
    ["第一主成分 PC1 代表資料中的哪個方向？", "最小變異方向", "隨機方向", "最大變異（最分散）方向", "純雜訊", 20, 3],
    ["PCA 的「分數（scores, T）」描述的是？", "每個樣本在新座標上的位置", "代謝物的組成", "模型誤差", "波長", 20, 1],
    ["「負荷量（loadings, P）」告訴我們？", "樣本數量", "模型準確率", "每個代謝物如何組成主成分", "樣本位置", 20, 3],
    ["陡坡圖（scree plot）主要用來決定？", "投影片顏色", "要保留幾個主成分", "樣本數", "品種名稱", 20, 2],
    ["本課 PCA 用 Hotelling T² 抓出 5 支離群酒，它們的共同點？", "都是假酒", "隨機出現", "同一個 2001 舊年份批次", "都是紅酒", 30, 3],
    ["關於 PCA，下列何者正確？", "需要品種標籤才能運作", "沒看過標籤，樣本卻依品種自動分群", "它會直接預測含量", "只能處理兩個變數", 20, 2],
]
KAHOOT_PLSDA = [
    ["PLS-DA 與 PCA 最大的差別是？", "PLS-DA 是監督式（用到類別標籤）", "兩者都不用標籤", "PLS-DA 一定較不準", "PLS-DA 不能畫圖", 20, 1],
    ["PLS-DA 找方向的準則是？", "最大化 X 的總變異", "隨機選方向", "最大化指紋投影與類別 y 的共變異", "最小化樣本數", 20, 3],
    ["PLS-DA 的「潛在變量（LV）」類似 PCA 的什麼？", "樣本", "標籤", "主成分", "誤差", 20, 3],
    ["為什麼要用「交叉驗證」評估模型？", "讓準確率看起來更高", "避免用訓練資料自我評分造成虛高", "純粹為了加快速度", "沒有必要", 20, 2],
    ["選 LV 時「過度擬合」的表現是？", "LV 越多交叉驗證一定越好", "與 LV 數無關", "LV 太多會開始硬背雜訊、CV 不再進步", "LV 越少一定越糟", 30, 3],
    ["本課 Chardonnay vs Sauvignon Blanc，只用 1 個 LV 的 CV 準確率約？", "50%", "0%", "約 97%", "無法計算", 20, 3],
    ["混淆矩陣「對角線」上的數字代表？", "分類正確的數量", "分類錯誤的數量", "樣本總數", "代謝物數量", 20, 1],
    ["VIP 分數用來找出？", "樣本位置", "投影片顏色", "樣本數量", "對分類貢獻最大的代謝物", 20, 4],
    ["本課 VIP 榜首（酒石酸、脯胺酸等）的特別之處？", "與酒完全無關", "是污染物", "是葡萄酒中真實且重要的成分", "隨機挑選", 30, 3],
    ["下列何者是 PLS-DA 在食品上的典型應用？", "產地溯源、品種鑑別、摻偽偵測", "加熱殺菌", "包裝印刷設計", "計算食物熱量", 20, 1],
]

# ── Wordwall Match：詞卡 ↔ 定義 ──
WORDWALL_MATCH = [
    ["質譜 MS", "依質荷比 m/z 偵測化合物的高靈敏方法"],
    ["GC-MS", "氣相層析串聯質譜，產生樣本×代謝物峰表"],
    ["代謝指紋", "一個樣本上百種代謝物的整體輪廓"],
    ["log 轉換", "壓縮跨數量級的峰高動態範圍"],
    ["autoscale", "逐代謝物標準化，使其等權重貢獻"],
    ["PCA", "非監督降維，找資料最大變異方向"],
    ["PC1", "資料中變異最大的方向"],
    ["分數 scores（T）", "樣本在主成分座標上的位置"],
    ["負荷量 loadings（P）", "各代謝物對主成分的貢獻"],
    ["陡坡圖 scree", "決定要保留幾個主成分的圖"],
    ["Hotelling T²", "偵測離群／異常樣本的統計量"],
    ["PLS-DA", "監督式判別，把軸對齊類別分界"],
    ["潛在變量 LV", "PLS 的壓縮方向，最大化與類別的共變異"],
    ["交叉驗證", "訓練與測試分開，誠實估計準確率"],
    ["過度擬合", "模型過度複雜、硬背雜訊"],
    ["VIP", "衡量代謝物對分類貢獻的指標"],
]

# ── Wordwall Quiz：[題目, 正解, 錯1, 錯2, 錯3] ──
WORDWALL_QUIZ = [
    ["PCA 是哪一種學習？", "非監督式", "監督式", "強化學習", "迴歸"],
    ["PLS-DA 是哪一種學習？", "監督式", "非監督式", "分群", "降維（不用標籤）"],
    ["PC1 代表的方向是？", "最大變異", "最小變異", "隨機方向", "與 y 垂直"],
    ["分數（T）描述的是？", "樣本位置", "代謝物組成", "模型誤差", "波長"],
    ["負荷量（P）描述的是？", "代謝物如何組成主成分", "樣本位置", "模型準確率", "樣本數"],
    ["log+autoscale 前處理的目的？", "壓縮範圍並讓代謝物等權重", "刪除小峰", "增加樣本數", "改變品種"],
    ["Hotelling T² 用來做什麼？", "偵測離群樣本", "預測濃度", "選顏色", "計算熱量"],
    ["PLS-DA 找方向的準則？", "最大化與類別的共變異", "最大化 X 總變異", "最小化樣本數", "隨機選取"],
    ["為何要交叉驗證？", "避免自我評分造成虛高", "讓準確率更高", "純為加快速度", "讓畫面美觀"],
    ["過度擬合的特徵？", "LV 太多、硬背雜訊", "LV 太少", "樣本太多", "與模型複雜度無關"],
    ["VIP 高的代謝物代表？", "對分類貢獻大", "對分類無關", "是雜訊", "是樣本編號"],
    ["PLS-DA 的食品應用？", "產地溯源／品種鑑別／摻偽偵測", "加熱殺菌", "包裝印刷", "熱量計算"],
]

# ── Wordwall Rank Order：一個排序活動 = 一個 CSV（見技能踩坑） ──
WORDWALL_SORT = [
    {"file": "wordwall_sort_pca.csv", "title": "PCA 分析流程（由先到後排序）",
     "items": ["GC-MS 取得代謝物峰表", "log + autoscale 前處理", "執行 PCA",
               "看陡坡圖選成分數", "分數圖看分群／異常", "負荷量回連化學"]},
    {"file": "wordwall_sort_plsda.csv", "title": "建立 PLS-DA 分類模型的步驟（由先到後）",
     "items": ["指定類別標籤 y", "log + autoscale 前處理", "擬合 PLS-DA 取得 LV",
               "交叉驗證選 LV 數", "看混淆矩陣評估", "用 VIP 找關鍵代謝物"]},
]

# ── Gimkit 快答 ──
GIMKIT_QA = [
    ["質譜依什麼偵測化合物？", "質荷比 m/z"], ["GC-MS 產生的資料形式？", "樣本×代謝物矩陣"],
    ["峰高取對數的目的？", "壓縮動態範圍"], ["autoscale 讓代謝物？", "等權重貢獻"],
    ["PCA 是監督還是非監督？", "非監督"], ["PC1 代表什麼方向？", "最大變異"],
    ["分數 T 看的是？", "樣本"], ["負荷量 P 看的是？", "變數（代謝物）"],
    ["決定保留幾個 PC 看什麼圖？", "陡坡圖 scree"], ["偵測離群樣本的統計量？", "Hotelling T²"],
    ["本課離群的是哪個年份批次？", "2001"], ["PCA 不用什麼就能分群？", "標籤"],
    ["PLS-DA 是監督還是非監督？", "監督"], ["PLS-DA 最大化什麼？", "指紋與類別的共變異"],
    ["PLS 的壓縮方向叫？", "潛在變量 LV"], ["訓練測試分開的評估法？", "交叉驗證"],
    ["LV 太多會造成？", "過度擬合"], ["本課二元分類的兩個品種？", "Chardonnay 與 Sauvignon Blanc"],
    ["1 個 LV 的 CV 準確率約？", "97%"], ["5 折 CV 的整體準確率？", "100%"],
    ["混淆矩陣對角線代表？", "正確分類數"], ["找關鍵代謝物用什麼分數？", "VIP"],
    ["VIP 榜首的酒酸？", "酒石酸"], ["葡萄酒含量最高的胺基酸？", "脯胺酸"],
    ["PLS-DA 三大食品應用？", "產地溯源、品種鑑別、摻偽偵測"], ["資料集來源研究編號？", "ST000006"],
]


def make_kahoot(rows, fname, title):
    wb = Workbook(); ws = wb.active; ws.title = "Kahoot"
    header = ["Question text", "Answer 1", "Answer 2", "Answer 3", "Answer 4", "Time limit (sec)", "Correct answer(s)"]
    hf = PatternFill("solid", fgColor="46178F"); hfont = Font(bold=True, color="FFFFFF", size=11)
    odd = PatternFill("solid", fgColor="F3EFF8"); even = PatternFill("solid", fgColor="FFFFFF")
    bd = Border(bottom=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"))
    ctr = Alignment("center", "center", wrap_text=True); lft = Alignment("left", "center", wrap_text=True)
    ws.column_dimensions["A"].width = 54
    for c in "BCDE": ws.column_dimensions[c].width = 24
    ws.column_dimensions["F"].width = 15; ws.column_dimensions["G"].width = 16
    for ci, h in enumerate(header, 1):
        c = ws.cell(1, ci, h); c.fill = hf; c.font = hfont; c.border = bd; c.alignment = ctr
    for ri, r in enumerate(rows, 2):
        fill = odd if ri % 2 else even
        for ci, (val, al) in enumerate(zip(r, [lft, lft, lft, lft, lft, ctr, ctr]), 1):
            c = ws.cell(ri, ci, val); c.fill = fill; c.border = bd; c.alignment = al
            if ci == 7: c.font = Font(bold=True, color="2ECC71")
    ws.freeze_panes = "A2"
    p = os.path.join(OUT, fname); wb.save(p)
    print(f"[OK] Kahoot {title}: {fname}  ({len(rows)} 題)")


def make_csv(fname, header, rows, label):
    p = os.path.join(OUT, fname)
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(header)
        for r in rows: w.writerow(r)
    print(f"[OK] {label}: {fname}  ({len(rows)} 列)")


if __name__ == "__main__":
    print("輸出資料夾:", OUT)
    make_kahoot(KAHOOT_PCA, "kahoot_pca.xlsx", "PCA")
    make_kahoot(KAHOOT_PLSDA, "kahoot_plsda.xlsx", "PLS-DA")
    make_csv("wordwall_match.csv", ["Term", "Definition"], WORDWALL_MATCH, "Wordwall Match")
    make_csv("wordwall_quiz.csv", ["Question", "Correct Answer", "Wrong Answer 1", "Wrong Answer 2", "Wrong Answer 3"], WORDWALL_QUIZ, "Wordwall Quiz")
    for s in WORDWALL_SORT:
        hdr = ["Question Title"] + [f"Item {i+1}" for i in range(len(s["items"]))]
        make_csv(s["file"], hdr, [[s["title"]] + s["items"]], "Wordwall Sort")
    make_csv("gimkit.csv", ["Question", "Answer"], GIMKIT_QA, "Gimkit")
    print("\n完成！Kahoot ×2、Wordwall Match/Quiz/Sort、Gimkit。")
